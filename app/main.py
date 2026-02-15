import re
import uuid
import zipfile
from datetime import date
from io import BytesIO
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from fastapi import Cookie, Depends, FastAPI, Form, Request, Response
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl.utils import get_column_letter, range_boundaries
from sqlalchemy.ext.asyncio import AsyncSession

from sqlalchemy import select

from app.database import close_db, get_db, init_db
from app.models import ExampleContent, Session as SessionModel
from app.services.gemini_service import GeminiService
from app.store import session_store

"""
Founder's Pilot アプリケーションのメインエントリポイントです。

FastAPIを使用したWebサーバーとして機能し、以下の役割を担います。
- 静的ファイルとHTMLテンプレートの配信
- クライアント（ブラウザ）からのチャットメッセージの受信とAI応答の返却
- セッション管理（Cookieを使用）
- 創業計画書の編集、保存、生成
- 完成した計画書（Excel）と記入例（PDF）のZIPダウンロード提供

主なエンドポイント:
- `/`: アプリケーションのルート（ダッシュボード）
- `/chat/*`: チャット機能関連（開始、メッセージ送信）
- `/plan/*`: 計画書の編集、保存、生成、ダウンロード
"""

# Load environment variables first
load_dotenv()

app = FastAPI(
    title="Founder's Pilot",
    description="""
    ## 創業計画書作成支援アプリケーション

    AIエージェントとの対話を通じて、創業計画書を作成するためのWebアプリケーションです。

    ### 主な機能
    - **AIチャット**: Gemini APIを使用した対話型ヒアリング
    - **計画書作成**: 10項目の創業計画書を段階的に作成
    - **編集機能**: 作成した計画書の編集・保存
    - **Excelエクスポート**: 日本政策金融公庫のテンプレート形式でダウンロード
    - **記入例提供**: 業種別の記入例PDFを自動選択して提供

    ### エンドポイント
    - `/`: ダッシュボード（HTMLページ）
    - `/chat/*`: チャット機能（AIとの対話）
    - `/plan/*`: 計画書の編集・保存・生成・ダウンロード
    """,
    version="1.0.0",
    contact={
        "name": "Founder's Pilot Support",
    },
)


# アプリケーション起動時の処理
@app.on_event("startup")
async def startup_event():
    """
    アプリケーション起動時にデータベースを初期化します。
    """
    await init_db()
    print("[INFO] Database initialized successfully")


# アプリケーション終了時の処理
@app.on_event("shutdown")
async def shutdown_event():
    """
    アプリケーション終了時にデータベース接続を閉じます。
    """
    await close_db()
    print("[INFO] Database connection closed")


# Initialize Gemini Service
gemini_service = GeminiService()

# --- 定数の定義 ---
# 大項目のフロー（ステッパー用）
ROADMAP_STEPS = [
    {"id": 1, "title": "構想・準備", "status": "current"},
    {"id": 2, "title": "創業計画書作成", "status": "upcoming"},
    {"id": 3, "title": "面談対策", "status": "upcoming"},
]

# 現在のフェーズ（創業計画書）のタスクリスト
TASKS = [
    {
        "id": "motivation",
        "title": "1. 創業の動機（創業されるのは、どのような目的、動機からですか）",
        "desc": "なぜこの事業を始めるのか、熱意と目的を言語化しましょう。",
        "status": "pending",
    },
    {
        "id": "background",
        "title": "2. 経営者の略歴等（略歴については、勤務先名だけではなく、担当業務や役職、身につけた技能等についても記載してください）",
        "desc": "これまでの経験が事業にどう活きるか整理します。",
        "status": "pending",
    },
    {
        "id": "service",
        "title": "3. 取扱商品・サービス",
        "desc": "商品・サービスの強みや特徴を明確にします。",
        "status": "pending",
    },
    {
        "id": "employees",
        "title": "4. 従業員",
        "desc": "常勤役員、従業員数などの体制を計画します。",
        "status": "pending",
    },
    {
        "id": "partners",
        "title": "5. 取引先・取引関係等",
        "desc": "販売先や仕入先、掛取引の条件などを整理します。",
        "status": "pending",
    },
    {
        "id": "related_companies",
        "title": "6. 関連企業（お申込人もしくは法人代表者または配偶者の方がご経営されている企業がある場合にご記入ください）",
        "desc": "関連する企業との関係性を整理します。",
        "status": "pending",
    },
    {
        "id": "loans",
        "title": "7. お借入の状況（法人の場合、代表者の方のお借入）",
        "desc": "個人的な借り入れや住宅ローンなどの状況を確認します。",
        "status": "pending",
    },
    {
        "id": "funds",
        "title": "8. 必要な資金と調達方法",
        "desc": "設備資金・運転資金の総額と、自己資金・借入金のバランスを計算します。",
        "status": "pending",
    },
    {
        "id": "outlook",
        "title": "9. 事業の見通し（月平均）",
        "desc": "創業当初と軌道に乗った後の売上・利益予測を立てます。",
        "status": "pending",
    },
    {
        "id": "free_description",
        "title": "10. 自由記述欄（アピールポイント、事業を行ううえでの悩み、希望するアドバイス等）",
        "desc": "事業のアピールポイントや悩み、希望するアドバイスなどを記載します。",
        "status": "pending",
    },
]

# Static files and Templates setup
BASE_DIR = Path(__file__).resolve().parent
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))


# --- ヘルパー関数 ---
def extract_sections_from_text(plan_text: str) -> dict:
    """
    創業計画書の全文テキストから各セクションを抽出します。

    Args:
        plan_text (str): 創業計画書の全文（マークダウン形式）

    Returns:
        dict: 各セクションの辞書 {"motivation": "...", "background": "...", ...}
    """
    heading_order = [
        ("motivation", "創業の動機"),
        ("background", "経営者の略歴等"),
        ("service", "取扱商品・サービス"),
        ("employees", "従業員"),
        ("partners", "取引先・取引関係等"),
        ("related_companies", "関連企業"),
        ("loans", "お借入の状況"),
        ("funds", "必要な資金と調達方法"),
        ("outlook", "事業の見通し"),
        ("free_description", "自由記述欄"),
    ]

    def build_section_pattern(current_label, next_labels):
        """見出しとその次の見出しまでの内容を抽出する正規表現パターンを生成"""
        next_part = "|".join(re.escape(label) for label in next_labels)
        if next_labels:
            return rf"^\s*#{{{0, 6}}}\s*\d*\.?\s*{re.escape(current_label)}\s*(?:\n|:|：)\s*(.*?)(?=^\s*#{{{0, 6}}}\s*\d*\.?\s*(?:{next_part}))"
        else:
            return rf"^\s*#{{{0, 6}}}\s*\d*\.?\s*{re.escape(current_label)}\s*(?:\n|:|：)\s*(.*?)$"

    def normalize_section_text(content: str) -> str:
        content = content.strip()
        content = re.sub(r"[\*#]+", "", content)
        content = re.sub(r"\n\s*\n+", "\n\n", content)
        return content.strip()

    sections = {}
    for idx, (key, label) in enumerate(heading_order):
        next_labels = [next_label for _, next_label in heading_order[idx + 1 :]]
        pattern = build_section_pattern(label, next_labels)
        match = re.search(pattern, plan_text, re.DOTALL | re.MULTILINE)
        if match:
            sections[key] = normalize_section_text(match.group(1))
        else:
            sections[key] = None

    return sections


def build_plan_text_from_sections(sections: dict) -> str:
    """
    各セクションから創業計画書の全文テキストを再構築します。

    Args:
        sections (dict): 各セクションの辞書

    Returns:
        str: 創業計画書の全文（マークダウン形式）
    """
    heading_order = [
        ("motivation", "1. 創業の動機"),
        ("background", "2. 経営者の略歴等"),
        ("service", "3. 取扱商品・サービス"),
        ("employees", "4. 従業員"),
        ("partners", "5. 取引先・取引関係等"),
        ("related_companies", "6. 関連企業"),
        ("loans", "7. お借入の状況"),
        ("funds", "8. 必要な資金と調達方法"),
        ("outlook", "9. 事業の見通し"),
        ("free_description", "10. 自由記述欄"),
    ]

    plan_text_parts = []
    for key, title in heading_order:
        content = sections.get(key)
        if content:
            plan_text_parts.append(f"## {title}\n\n{content}")

    return "\n\n".join(plan_text_parts)


def extract_single_section_from_response(ai_response: str, task_id: str) -> str | None:
    """
    AI応答から特定のタスクに対応するセクションの内容を抽出します。

    AI応答内の [[CONTENT_START]] と [[CONTENT_END]] マーカーで囲まれた
    コンテンツを取得します。

    Args:
        ai_response (str): AIからの応答テキスト
        task_id (str): タスクID（例: "motivation", "background"）

    Returns:
        str | None: 抽出されたセクション内容。見つからない場合はNone
    """
    # [[CONTENT_START]] と [[CONTENT_END]] の間のコンテンツを取得
    start_marker = "[[CONTENT_START]]"
    end_marker = "[[CONTENT_END]]"

    start_pos = ai_response.find(start_marker)
    end_pos = ai_response.find(end_marker)

    if start_pos == -1 or end_pos == -1:
        print(f"[WARNING] Content markers not found in AI response")
        print(f"[WARNING]   start_marker found: {start_pos != -1}")
        print(f"[WARNING]   end_marker found: {end_pos != -1}")
        return None

    if start_pos >= end_pos:
        print(f"[ERROR] Invalid marker positions: start={start_pos}, end={end_pos}")
        return None

    # マーカー間のコンテンツを取得（マーカー自体は除外）
    content_start = start_pos + len(start_marker)
    content = ai_response[content_start:end_pos]

    # 前後の空白・改行を削除（内部の改行は保持）
    content = content.strip()

    return content if content else None


# 業種・セクションの表示名定義
INDUSTRY_DISPLAY_NAMES = {
    "software": "ソフトウェア開発業（ITサービス、Webサービス、マッチングサービス、アプリ開発等）",
    "restaurant": "洋風居酒屋（飲食店）",
    "beauty": "美容業",
    "car_sales": "中古自動車販売業",
    "apparel": "婦人服・子供服小売業",
    "construction": "内装工事業",
    "cram_school": "学習塾",
    "dentist": "歯科診療所",
    "care_service": "介護サービス",
}

SECTION_DISPLAY_NAMES = {
    "motivation": "1. 創業の動機",
    "background": "2. 経営者の略歴等",
    "service": "3. 取扱商品・サービス",
    "employees": "4. 従業員",
    "partners": "5. 取引先・取引関係等",
    "related_companies": "6. 関連企業",
    "loans": "7. お借入の状況",
    "funds": "8. 必要な資金と調達方法",
    "outlook": "9. 事業の見通し（月平均）",
}

# 日本政策金融公庫「創業計画書セルフチェックリスト」(2023年版) より
SELF_CHECK_ITEMS: dict[str, list[str]] = {
    "motivation": [
        "創業への熱意や創業を志すまでの経緯を記載していますか？",
    ],
    "background": [
        "担当した業務や役職、実績などを記載していますか？",
        "身に着けた資格・スキルなどがあれば、それらについて記載していますか？",
    ],
    "service": [
        "誰に、何を、いくらで販売するか記載していますか？",
        "商品・サービスのセールスポイントを記載していますか？",
        "販売ターゲットに合った販売戦略について記載していますか？",
        "競合他社や市場について調べて、記載していますか？",
    ],
    "partners": [
        "入金や支払いのタイミングなど、取引形態を記載していますか？",
    ],
    "funds": [
        "見積金額が適切か、相場を調べたり相見積もりを取得するなどして検証していますか？",
        "事業開始後の運転資金（半年程度の赤字補てん資金など）について検討していますか？",
        "自己資金が少なく、借入依存の資金調達計画になっていませんか？",
    ],
    "outlook": [
        "計算根拠をもって売上高や売上原価の予測を立てていますか？",
        "経費に漏れがないか検討していますか？",
        "利益から借入の返済が可能な収支計画となっていますか？",
    ],
}

SECTION_ORDER = [
    "motivation",
    "background",
    "service",
    "employees",
    "partners",
    "related_companies",
    "loans",
    "funds",
    "outlook",
]

INDUSTRY_ORDER = [
    "software",
    "restaurant",
    "beauty",
    "car_sales",
    "apparel",
    "construction",
    "cram_school",
    "dentist",
    "care_service",
]


def convert_western_to_wareki(text: str) -> str:
    """
    テキスト中の西暦年月表記（例: 2010年4月）を和暦（元号）表記に変換する。
    既に和暦表記になっている箇所は変換しない。

    元号の境界:
    - 昭和: 1926〜1989/1/7
    - 平成: 1989/1/8〜2019/4/30
    - 令和: 2019/5/1〜
    """

    def year_month_to_wareki(year: int, month: int | None) -> str:
        if year > 2019 or (year == 2019 and month is not None and month >= 5):
            wareki_year = year - 2018
            era = "令和"
        elif year == 2019:
            # 2019年1〜4月 → 平成31年
            wareki_year = 31
            era = "平成"
        elif year > 1989 or (year == 1989 and month is not None and month >= 2):
            wareki_year = year - 1988
            era = "平成"
        elif year == 1989:
            # 1989年1月 → 昭和64年
            wareki_year = 64
            era = "昭和"
        elif year >= 1926:
            wareki_year = year - 1925
            era = "昭和"
        else:
            return f"{year}年"  # 大正以前はそのまま

        year_str = "元" if wareki_year == 1 else str(wareki_year)
        return f"{era}{year_str}年"

    def replace_match(m: re.Match) -> str:
        year = int(m.group(1))
        month_str = m.group(2)  # "4月" or None
        month = int(m.group(3)) if m.group(3) else None
        wareki = year_month_to_wareki(year, month)
        return wareki + (month_str or "")

    # 西暦4桁（1900〜2099）の年月表記を対象にする。直前が数字の場合は除外
    pattern = r"(?<!\d)((?:19|20)\d{2})年((\d{1,2})月)?"
    return re.sub(pattern, replace_match, text)


async def build_examples_text(db: AsyncSession) -> str:
    """
    example_contentsテーブルから全業種の記入例を取得し、
    システムプロンプト用のテキストに整形して返します。
    """
    result = await db.execute(select(ExampleContent))
    examples = result.scalars().all()

    # 業種・セクションでインデックス化
    examples_dict: dict[str, dict[str, str]] = {}
    for ex in examples:
        examples_dict.setdefault(ex.industry_type, {})[ex.section_key] = ex.example_text

    parts = []
    for industry in INDUSTRY_ORDER:
        if industry not in examples_dict:
            continue
        display_name = INDUSTRY_DISPLAY_NAMES.get(industry, industry)
        parts.append(f"■{display_name}")
        parts.append("")
        for section_key in SECTION_ORDER:
            if section_key not in examples_dict[industry]:
                continue
            section_name = SECTION_DISPLAY_NAMES.get(section_key, section_key)
            parts.append(section_name)
            parts.append(examples_dict[industry][section_key])
            parts.append("")

    return "\n".join(parts)


async def get_section_example(
    db: AsyncSession, industry_type: str, section_key: str
) -> str | None:
    """
    指定した業種・セクションの記入例をDBから取得して返します。
    見つからない場合はNoneを返します。
    """
    if not industry_type or not section_key:
        return None
    result = await db.execute(
        select(ExampleContent).where(
            ExampleContent.industry_type == industry_type,
            ExampleContent.section_key == section_key,
        )
    )
    example = result.scalar_one_or_none()
    return example.example_text if example else None


@app.get("/", response_class=HTMLResponse)
async def read_root(
    request: Request,
    session_id: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    """
    アプリケーションのメインページ（ダッシュボード）を表示します。

    CookieからセッションIDを取得し、セッションが存在する場合は
    これまでのタスク進捗状況とチャット履歴を復元して表示します。

    Args:
        request (Request): FastAPIのリクエストオブジェクト
        session_id (str | None): クライアントのCookieから取得したセッションID
        db (AsyncSession): データベースセッション

    Returns:
        TemplateResponse: レンダリングされた `index.html`
    """
    # セッションIDがあればデータをロード
    user_tasks = TASKS  # デフォルト
    chat_messages_html = ""  # ここで履歴HTML文字列を作る

    if session_id:
        data = await session_store.load_session(db, session_id)
        if data:
            # データベースから取得したタスク状態をTASKSテンプレートにマージ
            task_states = data.get("task_states", {})
            user_tasks = []
            for task in TASKS:
                task_copy = task.copy()
                # データベースに保存されている状態があればそれを使用
                if task["id"] in task_states:
                    task_copy["status"] = task_states[task["id"]]
                user_tasks.append(task_copy)

            history = data.get("chat_history", [])

            # Gemini履歴を復元
            if history:
                gemini_service.restore_chat_session(session_id, history)

            # 表示用の履歴HTMLを再構築
            # ここでは簡易的に、historyの構造から role=user or model を判定してテンプレート適用
            # 本来は templates/components/message.html を再利用して文字列結合する
            for msg in history:
                role = msg["role"]
                # model -> "text", user -> "text"
                text_content = ""
                if "parts" in msg:
                    for part in msg["parts"]:
                        if "text" in part:
                            text_content += part["text"]

                is_user = role == "user"

                msg_html = templates.get_template("components/message.html").render(
                    {"request": request, "message": text_content, "is_user": is_user}
                )
                chat_messages_html += msg_html

    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "steps": ROADMAP_STEPS,
            "tasks": user_tasks,
            "initial_chat_history": chat_messages_html,
        },
    )


@app.get("/chat/start", response_class=HTMLResponse)
async def start_chat(
    request: Request,
    task_id: str | None = None,
    db: AsyncSession = Depends(get_db),
):
    """
    新しいチャットセッションを開始します。

    新規にUUIDを生成してセッションIDとし、AIエージェントからの初期挨拶メッセージを取得して返します。
    特定のタスクIDが指定された場合、そのタスクのヒアリングから開始するようにAIに指示します。

    Args:
        request (Request): FastAPIのリクエストオブジェクト
        task_id (str | None): 個別のタスクから開始する場合のタスクID（例: "motivation"）
        db (AsyncSession): データベースセッション

    Returns:
        TemplateResponse: 初期メッセージを含む `components/chat_interface.html`
    """
    session_id = str(uuid.uuid4())

    # データベースにセッションレコードを作成
    new_session = SessionModel(id=session_id)
    db.add(new_session)
    await db.commit()

    initial_task_title = None
    if task_id:
        # TASKSからタイトルを検索
        for task in TASKS:
            if task["id"] == task_id:
                initial_task_title = task["title"]
                break

    # DBから記入例テキストを構築してシステムプロンプトに渡す
    examples_text = await build_examples_text(db)

    # エージェントからの初期挨拶を生成
    # GeminiService.start_chat_session requires session_id
    tGreeting = await gemini_service.start_chat_session(
        session_id, initial_task_title, examples_text
    )

    # 業種選択マーカーの検出
    industry_selector_html = ""
    if "[[INDUSTRY_SELECTOR]]" in tGreeting:
        tGreeting = tGreeting.replace("[[INDUSTRY_SELECTOR]]", "")
        # 業種選択ボタンのHTMLを生成
        industry_selector_html = templates.get_template(
            "components/industry_selector.html"
        ).render({"request": request})

    # 完全なチャットインターフェースを返す（chat_interface.html テンプレートを使用）
    response = templates.TemplateResponse(
        "components/chat_interface.html",
        {
            "request": request,
            "message": tGreeting,
            "is_user": False,
            "additional_content": industry_selector_html,  # 業種選択ボタンを追加コンテンツとして渡す
        },
    )
    # 簡易的にCookieでセッションIDを管理
    response.set_cookie(key="session_id", value=session_id)
    return response


@app.post("/chat/approve_draft", response_class=HTMLResponse)
async def approve_draft(
    request: Request,
    task_id: str = Form(...),
    draft_content: str = Form(default=None),  # 表示時に変換済みのドラフト内容
    skip_verification: bool = Form(default=False),  # 記入例検証をスキップするフラグ
    session_id: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    """
    ドラフト承認ボタン（OK）がクリックされた際の処理。

    Geminiに依存せず、確実に次のステップに進みます。

    処理フロー:
    1. 直前のAI応答からドラフト内容を抽出
    2. タスクを完了状態に更新
    3. データベースに保存
    4. 次のタスクの質問を生成
    """
    if not session_id:
        return HTMLResponse("Session not found", status_code=400)

    # セッションデータをロード
    import copy

    current_tasks = copy.deepcopy(TASKS)
    session_data = await session_store.load_session(db, session_id)

    if session_data and "task_states" in session_data:
        task_states = session_data["task_states"]
        current_tasks = []
        for task in TASKS:
            task_copy = task.copy()
            if task["id"] in task_states:
                task_copy["status"] = task_states[task["id"]]
            current_tasks.append(task_copy)

    # チャット履歴から直前のAI応答を取得
    history_data = gemini_service.get_chat_history(session_id)
    last_ai_response = None
    if history_data:
        # 最後のAIメッセージを探す
        for msg in reversed(history_data):
            if msg.get("role") == "model":
                last_ai_response = msg.get("parts", [{}])[0].get("text", "")
                break

    # ドラフト内容を取得
    # 優先: フォームから渡された表示済み内容を使用
    if draft_content and draft_content.strip():
        draft_content = draft_content.strip()
    else:
        # フォールバック: チャット履歴からドラフト内容を抽出
        print("[WARNING] No draft_content from form, extracting from chat history")
        draft_content = None
        if last_ai_response:
            draft_content = extract_single_section_from_response(
                last_ai_response, task_id
            )
        if not draft_content and last_ai_response:
            draft_content = (
                last_ai_response.replace("[[CONTENT_START]]", "")
                .replace("[[CONTENT_END]]", "")
                .replace("[[DRAFT_PROPOSED]]", "")
                .strip()
            )

    # backgroundセクションの場合は必ず和暦変換を適用（サーバーサイドで確実に処理）
    if draft_content and task_id == "background":
        draft_content = convert_western_to_wareki(draft_content)

    # --- 記入例との比較検証 ---
    # 前回の検証でフィードバックを表示済みかどうかをCookieで判定
    verification_flagged = request.cookies.get(f"vf_{task_id}")
    industry_type_for_verify = (
        session_data.get("industry_type") if session_data else None
    )
    if not skip_verification and draft_content and industry_type_for_verify:
        current_example = await get_section_example(
            db, industry_type_for_verify, task_id
        )
        if current_example:
            import html as _html

            escaped_draft = _html.escape(draft_content, quote=True)
            section_label = SECTION_DISPLAY_NAMES.get(task_id, task_id)
            section_verification = await gemini_service.verify_section_draft(
                section_label,
                draft_content,
                current_example,
                self_check_items=SELF_CHECK_ITEMS.get(task_id),
            )
            if section_verification.get("has_issues"):
                # 問題あり → フィードバック + 「このままOKにする」ボタン
                # Cookie フラグを立てて「ブラッシュアップ中」と記憶する
                feedback_text = section_verification.get("feedback", "")
                # AIがMarkdownを返した場合に備えてHTMLタグに変換
                feedback_text = re.sub(
                    r"\*\*(.+?)\*\*", r"<strong>\1</strong>", feedback_text
                )
                early_user_msg_html = templates.get_template(
                    "components/message.html"
                ).render(
                    {"request": request, "message": "この内容でOKです", "is_user": True}
                )
                feedback_msg = (
                    f"記入例と比較したところ、改善できる点が見つかりました。\n\n"
                    f"{feedback_text}\n\n"
                    f"改善したい場合はチャットでご相談ください。"
                    f"このまま進める場合は下のボタンを押してください。"
                )
                ai_feedback_html = templates.get_template(
                    "components/message.html"
                ).render(
                    {"request": request, "message": feedback_msg, "is_user": False}
                )
                force_ok_button = f"""<div class="flex justify-start ml-12 mb-4">
  <form>
    <input type="hidden" name="task_id" value="{task_id}">
    <input type="hidden" name="draft_content" value="{escaped_draft}">
    <input type="hidden" name="skip_verification" value="true">
    <button hx-post="/chat/approve_draft"
            hx-target="#chat-history"
            hx-swap="beforeend"
            hx-include="closest form"
            class="px-4 py-2 bg-slate-500 text-white text-sm rounded hover:bg-slate-600 shadow-sm font-medium">
      このままOKにする →
    </button>
  </form>
</div>"""
                resp = HTMLResponse(
                    content=early_user_msg_html + ai_feedback_html + force_ok_button
                )
                resp.set_cookie(key="session_id", value=session_id)
                # ブラッシュアップ中フラグをCookieにセット（1時間有効）
                resp.set_cookie(
                    key=f"vf_{task_id}", value="1", max_age=3600, httponly=True
                )
                return resp
            elif verification_flagged:
                # ブラッシュアップ後に再検証してOKだった場合でも自動進行しない
                # → 明示的な「次のセクションへ」ボタンを表示して確認を求める
                early_user_msg_html = templates.get_template(
                    "components/message.html"
                ).render(
                    {"request": request, "message": "この内容でOKです", "is_user": True}
                )
                ok_msg = "内容が改善されました。記入例と比較して問題は見つかりませんでした。\n\nこのまま次のセクションへ進む場合は下のボタンを押してください。"
                ai_ok_html = templates.get_template("components/message.html").render(
                    {"request": request, "message": ok_msg, "is_user": False}
                )
                advance_button = f"""<div class="flex justify-start ml-12 mb-4">
  <form>
    <input type="hidden" name="task_id" value="{task_id}">
    <input type="hidden" name="draft_content" value="{escaped_draft}">
    <input type="hidden" name="skip_verification" value="true">
    <button hx-post="/chat/approve_draft"
            hx-target="#chat-history"
            hx-swap="beforeend"
            hx-include="closest form"
            class="px-4 py-2 bg-green-600 text-white text-sm rounded hover:bg-green-700 shadow-sm font-medium">
      次のセクションへ進む →
    </button>
  </form>
</div>"""
                resp = HTMLResponse(
                    content=early_user_msg_html + ai_ok_html + advance_button
                )
                resp.set_cookie(key="session_id", value=session_id)
                # ブラッシュアップフラグを削除（次のセクションへ進む前にクリア）
                resp.delete_cookie(key=f"vf_{task_id}")
                return resp

    # タスクを完了状態に更新
    task_found = False
    for task in current_tasks:
        if task["id"] == task_id:
            task["status"] = "done"
            task_found = True
            break

    if not task_found:
        print(f"[ERROR] Task {task_id} not found in task list")
        return HTMLResponse("Invalid task ID", status_code=400)

    # ドラフト内容をデータベースに保存
    if draft_content:
        section_update = {task_id: draft_content}
        await session_store.save_session(
            db,
            session_id,
            current_tasks,
            history_data,
            sections=section_update,
        )

    # 次のタスクを見つける
    next_task = None
    for i, task in enumerate(TASKS):
        if task["id"] == task_id:
            if i + 1 < len(TASKS):
                next_task = TASKS[i + 1]
            break

    # 応答メッセージを生成
    user_msg_html = templates.get_template("components/message.html").render(
        {"request": request, "message": "この内容でOKです", "is_user": True}
    )

    # 次のセクションの記入例をDBから取得
    # ※AIへのメッセージにも含めて確実に参照させる
    industry_type = session_data.get("industry_type") if session_data else None
    next_example_text = None
    # 次のセクションをAIに明示して指定（AIが独自に順番を判断してスキップするのを防ぐ）
    if next_task:
        next_section_name = SECTION_DISPLAY_NAMES.get(
            next_task["id"], next_task["title"]
        )
        ai_ok_message = f"この内容でOKです。次のセクションは「{next_section_name}」です。このセクションについてヒアリングを開始してください。"

        # 記入例を取得してAIへのメッセージに含める
        if industry_type:
            next_example_text = await get_section_example(
                db, industry_type, next_task["id"]
            )
            if next_example_text:
                ai_ok_message += f"\n\n【このセクションの記入例（参考）】\n{next_example_text}\n\n記入例は文体・粒度の参考にしてください。記入例をそのまま返答に含めてはいけません。"
    else:
        ai_ok_message = "この内容でOKです。すべてのセクションが完了しました。"

    next_example_html = ""
    if next_task and industry_type and next_example_text:
        section_name = SECTION_DISPLAY_NAMES.get(next_task["id"], next_task["title"])
        industry_label = INDUSTRY_DISPLAY_NAMES.get(industry_type, industry_type)
        import html as html_module

        escaped_example = html_module.escape(next_example_text).replace("\n", "<br>")
        next_example_html = f"""<div class="flex justify-start my-2">
  <div class="bg-amber-50 border border-amber-200 px-4 py-3 rounded-lg max-w-2xl text-sm ml-12">
    <div class="font-semibold text-amber-800 mb-2">📋 {section_name} の記入例（{industry_label}の場合）</div>
    <div class="text-gray-700 text-xs leading-relaxed">{escaped_example}</div>
  </div>
</div>"""
    # 全セクション完了の場合はAI呼び出しをスキップ（AIが[[DRAFT_PROPOSED]]を含む余計な応答を返すのを防ぐ）
    if next_task:
        # 次のセクションがある場合のみAIを呼び出して次の質問を生成させる
        ai_next_response = await gemini_service.generate_response(
            session_id, ai_ok_message
        )

        # マーカーとコンテンツの削除処理
        # [[CONTENT_START]] ~ [[CONTENT_END]] の部分を削除（承認済みコンテンツは再表示不要）
        ai_next_response = re.sub(
            r"\[\[CONTENT_START\]\].*?\[\[CONTENT_END\]\]",
            "",
            ai_next_response,
            flags=re.DOTALL,
        )

        # [[COMPLETED:xxx]] マーカーを削除
        ai_next_response = re.sub(r"\[\[COMPLETED:[a-z_]+\]\]", "", ai_next_response)
        # [[DRAFT_PROPOSED]] マーカーを削除（完了メッセージに誤って含まれる場合）
        ai_next_response = re.sub(r"\[\[DRAFT_PROPOSED\]\]", "", ai_next_response)

        # 過剰な空白・改行を整理
        ai_next_response = re.sub(r"\n{3,}", "\n\n", ai_next_response)
        ai_next_response = ai_next_response.strip()

        # フォールバック: Geminiが適切な応答を返さない場合
        if not ai_next_response or len(ai_next_response.strip()) < 20:
            ai_next_response = f"承認ありがとうございます。\n\nそれでは次に、「{next_task['title']}」についてお伺いします。\n\n{next_task['desc']}\n\nどのような内容でしょうか？"

        ai_msg_html = templates.get_template("components/message.html").render(
            {"request": request, "message": ai_next_response, "is_user": False}
        )
    else:
        # 全セクション完了時はAIを呼ばずメッセージも表示しない
        ai_msg_html = ""

    # チェックボックス更新用のOOB HTML
    task_update_html = f"""
    <input id="task-{task_id}"
           name="task-{task_id}"
           type="checkbox"
           class="h-4 w-4 rounded border-gray-300 text-blue-600 focus:ring-blue-600"
           checked
           hx-swap-oob="true">
    """

    # 全セクション完了時の案内バナー
    completion_banner_html = ""
    if not next_task:
        completion_banner_html = """
<div class="flex justify-center my-4">
  <div class="bg-indigo-50 border border-indigo-300 px-6 py-4 rounded-lg max-w-xl text-center shadow-sm">
    <div class="text-indigo-700 font-bold text-base mb-1">🎉 全セクションの入力が完了しました！</div>
    <div class="text-gray-600 text-sm">左側の <span class="font-semibold text-indigo-600">「創業計画書案を作成する」</span> ボタンをクリックしてください。</div>
  </div>
</div>
"""

    response = HTMLResponse(
        content=user_msg_html
        + next_example_html
        + ai_msg_html
        + completion_banner_html
        + task_update_html
    )
    response.set_cookie(key="session_id", value=session_id)
    # ブラッシュアップフラグが残っている場合はクリア（このセクションは正常に完了）
    response.delete_cookie(key=f"vf_{task_id}")

    return response


@app.post("/chat/select_industry", response_class=HTMLResponse)
async def select_industry(
    request: Request,
    industry: str = Form(...),
    session_id: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    """
    業種選択ボタンがクリックされた際の処理。

    業種を保存し、AIの応答を含む完全なチャットインターフェースを返します。
    """
    if not session_id:
        return HTMLResponse("Session not found", status_code=400)

    # 業種タイプのマッピング（番号 → 内部キー）
    industry_mapping = {
        "1": "restaurant",
        "2": "beauty",
        "3": "car_sales",
        "4": "apparel",
        "5": "software",
        "6": "construction",
        "7": "cram_school",
        "8": "dentist",
        "9": "care_service",
    }

    # AIに伝える業種名ラベル（番号 → 日本語表示名）
    industry_label_mapping = {
        "1": "洋風居酒屋（飲食店）",
        "2": "美容業",
        "3": "中古自動車販売業",
        "4": "婦人服・子供服小売業",
        "5": "ソフトウェア開発業（ITサービス、Webサービス、マッチングサービス、アプリ開発等）",
        "6": "内装工事業",
        "7": "学習塾",
        "8": "歯科診療所",
        "9": "介護サービス",
    }

    industry_type = industry_mapping.get(industry)
    if not industry_type:
        return HTMLResponse("Invalid industry selection", status_code=400)

    # データベースに業種タイプを保存
    await session_store.update_industry_type(db, session_id, industry_type)

    # AIに業種名を明示して伝える（数字だけだと誤判定する恐れがあるため）
    industry_label = industry_label_mapping.get(industry, industry)
    ai_message = f"「{industry_label}」を選択しました。"

    # 最初のセクション（創業の動機）の記入例をDBから取得
    # ※AIへのメッセージには含めず、テンプレートに渡してHTMLとして直接表示する
    #   （AIはシステムプロンプトで全記入例を把握済み）
    first_example = await get_section_example(db, industry_type, "motivation")
    section_name = SECTION_DISPLAY_NAMES.get("motivation", "1. 創業の動機")
    example_label = (
        f"{section_name} の記入例（{industry_label}の場合）" if first_example else None
    )

    ai_response_text = await gemini_service.generate_response(session_id, ai_message)

    # チャット履歴を保存
    history_data = gemini_service.get_chat_history(session_id)
    await session_store.save_session(db, session_id, [], history_data)

    # 完全なチャットインターフェースを返す
    return templates.TemplateResponse(
        "components/chat_interface.html",
        {
            "request": request,
            "message": ai_response_text,
            "is_user": False,
            "show_industry_selection": f"{industry}",  # 選択された業種番号をユーザーメッセージとして表示
            "example_text": first_example,
            "example_label": example_label,
        },
    )


@app.post("/chat/message", response_class=HTMLResponse)
async def chat_message(
    request: Request,
    user_message: str = Form(...),
    session_id: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    """
    ユーザーからのチャットメッセージを処理し、AIからの応答を返します。

    主な処理フロー:
    1. ユーザーメッセージをチャット履歴に追加
    2. Gemini Service を用いてAI応答を生成
    3. AI応答内の特殊コマンド（マーカー）を検出し、以下の処理を実行
       - `[[DRAFT_PROPOSED]]`: ドラフト提案ボタン（OK/修正）を表示
       - `[[COMPLETED:{task_id}]]`: 指定されたタスクを完了状態(done)に更新し、チェックボックスをオンにする
    4. 更新されたセッション状態（タスク、履歴）を保存

    Args:
        request (Request): FastAPIのリクエストオブジェクト
        user_message (str): フォームから送信されたユーザーのメッセージ
        session_id (str | None): CookieセッションID
        db (AsyncSession): データベースセッション

    Returns:
        HTMLResponse: ユーザーメッセージ、AIメッセージ、およびもしあればUI更新用HTML（OOB Swap）を含むフラグメント
    """
    # ユーザーのメッセージを表示するためのHTML
    user_msg_html = templates.get_template("components/message.html").render(
        {"request": request, "message": user_message, "is_user": True}
    )

    # セッションIDがない場合のフォールバック（途中でCookieが消えた場合など）
    # 本来はエラーにするか再接続を促すべきだが、ここでは新規作成して続行
    if not session_id:
        session_id = str(uuid.uuid4())

    # AI応答の生成
    ai_response_text = await gemini_service.generate_response(session_id, user_message)

    # 業種選択の検出と処理
    industry_mapping = {
        "1": "restaurant",
        "2": "beauty",
        "3": "car_sales",
        "4": "apparel",
        "5": "software",
        "6": "construction",
        "7": "cram_school",
        "8": "dentist",
        "9": "care_service",
    }

    # ユーザーメッセージから業種番号を検出
    user_message_stripped = user_message.strip()
    if user_message_stripped in industry_mapping:
        industry_type = industry_mapping[user_message_stripped]
        await session_store.update_industry_type(db, session_id, industry_type)

    # タスク完了マーカーの検出と処理
    task_update_html = ""
    draft_buttons_html = ""
    industry_selector_html = ""

    # 業種選択マーカーの検出
    if "[[INDUSTRY_SELECTOR]]" in ai_response_text:
        ai_response_text = ai_response_text.replace("[[INDUSTRY_SELECTOR]]", "")
        # 業種選択ボタンのHTMLを生成
        industry_selector_html = templates.get_template(
            "components/industry_selector.html"
        ).render({"request": request})

    # セッションからタスク状態をロード（deep copy）
    import copy

    current_tasks = copy.deepcopy(TASKS)  # デフォルト値
    session_data = await session_store.load_session(db, session_id)
    if session_data and "task_states" in session_data:
        # データベースから取得したタスク状態をTASKSテンプレートにマージ
        task_states = session_data["task_states"]
        current_tasks = []
        for task in TASKS:
            task_copy = task.copy()
            # データベースに保存されている状態があればそれを使用
            if task["id"] in task_states:
                task_copy["status"] = task_states[task["id"]]
            current_tasks.append(task_copy)

    # ドラフト提示マーカーの検出（フォールバック検出を含む）
    # [[DRAFT_PROPOSED]]がなくても、ドラフト的な内容を含む場合はOKボタンを表示する
    DRAFT_PHRASES = [
        "でよろしいでしょうか",
        "いかがでしょうか",
        "ご確認ください",
        "以下の内容で",
    ]
    has_draft_proposed = "[[DRAFT_PROPOSED]]" in ai_response_text
    if not has_draft_proposed:
        if "[[CONTENT_START]]" in ai_response_text:
            has_draft_proposed = True
        elif any(phrase in ai_response_text for phrase in DRAFT_PHRASES):
            has_draft_proposed = True

    if has_draft_proposed:
        ai_response_text = ai_response_text.replace("[[DRAFT_PROPOSED]]", "")

        # 現在進行中のタスク（最初のpendingタスク）を特定
        current_task_id = None
        for task in current_tasks:
            if task["status"] == "pending":
                current_task_id = task["id"]
                break

        if current_task_id:
            # backgroundセクション（経営者の略歴等）のドラフト表示時に西暦→和暦変換
            if current_task_id == "background":
                ai_response_text = convert_western_to_wareki(ai_response_text)

            # 表示済み（変換済み）のドラフト内容を抽出してhidden inputに埋め込む
            # → approve_draftがそのままDBに保存するため、二重変換不要
            import html as _html

            content_match = re.search(
                r"\[\[CONTENT_START\]\](.*?)\[\[CONTENT_END\]\]",
                ai_response_text,
                re.DOTALL,
            )
            pending_draft = content_match.group(1).strip() if content_match else ""
            escaped_draft = _html.escape(pending_draft, quote=True)

            draft_buttons_html = f"""
            <form>
                <input type="hidden" name="task_id" value="{current_task_id}">
                <input type="hidden" name="draft_content" value="{escaped_draft}">
                <div class="flex gap-4 mt-2 mb-4 ml-12">
                    <button type="button"
                            hx-post="/chat/approve_draft"
                            hx-include="closest form"
                            hx-target="#chat-history"
                            hx-swap="beforeend"
                            class="px-4 py-2 bg-blue-600 text-white rounded-lg text-sm hover:bg-blue-700 font-bold transition-colors">
                        OK(次の項目へ)
                    </button>
                    <button type="button"
                            hx-post="/chat/message"
                            hx-vals='{{"user_message": "文面を修正したいです"}}'
                            hx-target="#chat-history"
                            hx-swap="beforeend"
                            class="px-4 py-2 bg-white border border-gray-300 text-gray-700 rounded-lg text-sm hover:bg-gray-50 transition-colors">
                        文面を修正する
                    </button>
                </div>
            </form>
            """
        else:
            print("[WARNING] No pending task found for draft buttons")

    # コンテンツマーカーを削除（表示用。ドラフト内容はOKボタン経由で保存するため表示不要）
    ai_response_text = ai_response_text.replace("[[CONTENT_START]]", "")
    ai_response_text = ai_response_text.replace("[[CONTENT_END]]", "")

    # [[COMPLETED:xxx]] マーカーをAIが誤出力した場合のクリーンアップ（保存処理は行わない）
    ai_response_text = re.sub(
        r"\n*\[\[COMPLETED:[a-z_]+\]\]\n*", "\n\n", ai_response_text
    )

    # 過剰な改行を整理
    ai_response_text = re.sub(r"\n{3,}", "\n\n", ai_response_text)

    ai_msg_html = templates.get_template("components/message.html").render(
        {"request": request, "message": ai_response_text, "is_user": False}
    )

    response = HTMLResponse(
        content=user_msg_html
        + ai_msg_html
        + industry_selector_html
        + draft_buttons_html
        + task_update_html
    )

    # セッションIDを再設定（有効期限延長などの効果もあるが、とりあえず念の為）
    response.set_cookie(key="session_id", value=session_id)

    # --- セッション状態の保存 ---
    # タスク完了時に既にセクションとタスク状態を保存している場合があるため、
    # ここでは主にチャット履歴の最終更新を行う
    # 1. GeminiServiceから現在のチャット履歴を取得
    history_data = gemini_service.get_chat_history(session_id)
    # 2. タスク状態とチャット履歴を保存（セクションは既に保存済みの場合はスキップ）
    await session_store.save_session(db, session_id, current_tasks, history_data)

    return response


@app.get(
    "/plan/edit",
    response_class=HTMLResponse,
    summary="計画書編集画面の取得",
    description="""
    創業計画書の編集画面（エディタ）を取得します。

    ⚠️ **Swagger UIでの直接テストはできません**
    - このエンドポイントはCookieによるセッション管理が必要です
    - Swagger UIではCookieパラメータを正しく送信できません

    **テスト方法:**
    - 下記の「Try it out」→「session_id」を入力→「Execute」で表示されるCurlコマンドをコピー
    - ターミナルで実行
    """,
)
async def edit_plan(
    request: Request,
    session_id: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    """
    創業計画書の編集画面（エディタ）を取得します。

    セッションに保存されている現在の計画書セクションを読み込み、
    全文として結合してから編集用フォームにセットして返します。

    Args:
        request (Request): FastAPIのリクエストオブジェクト
        session_id (str | None): CookieセッションID
        db (AsyncSession): データベースセッション

    Returns:
        TemplateResponse: `components/plan_editor.html`
    """
    if not session_id:
        return HTMLResponse("Session not found", status_code=400)

    data = await session_store.load_session(db, session_id)

    sections = {}
    if data and data.get("sections"):
        sections = data["sections"]
    else:
        # 空のセクション辞書を用意（テンプレートでエラーを防ぐ）
        sections = {
            "motivation": None,
            "background": None,
            "service": None,
            "employees": None,
            "partners": None,
            "related_companies": None,
            "loans": None,
            "funds": None,
            "outlook": None,
            "free_description": None,
        }
    # 業種に応じた記入例をDBから取得
    industry_type = data.get("industry_type") if data else None
    examples = {}
    if industry_type:
        for section_key in SECTION_ORDER:
            example = await get_section_example(db, industry_type, section_key)
            if example:
                examples[section_key] = example

    return templates.TemplateResponse(
        "components/plan_editor.html",
        {
            "request": request,
            "sections": sections,
            "examples": examples,
            "self_check_items": SELF_CHECK_ITEMS,
        },
    )


@app.post(
    "/plan/save",
    response_class=HTMLResponse,
    summary="計画書の保存",
    description="""
    編集された創業計画書テキストを保存します。

    ⚠️ **Swagger UIでの直接テストはできません**
    - このエンドポイントはCookieによるセッション管理が必要です
    - Swagger UIではCookieパラメータを正しく送信できません

    **テスト方法:**
    - 下記の「Try it out」→各フィールドと「session_id」を入力→「Execute」で表示されるCurlコマンドをコピー
    - ターミナルで実行
    """,
)
async def save_plan(
    request: Request,
    motivation: str = Form(default=""),
    background: str = Form(default=""),
    service: str = Form(default=""),
    employees: str = Form(default=""),
    partners: str = Form(default=""),
    related_companies: str = Form(default=""),
    loans: str = Form(default=""),
    funds: str = Form(default=""),
    outlook: str = Form(default=""),
    free_description: str = Form(default=""),
    session_id: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    """
    編集された創業計画書テキストを保存します。

    各セクションのフォームデータを受け取り、データベースに保存し、
    保存後は閲覧モード（Viewer）のHTMLを返します。

    Args:
        request (Request): FastAPIのリクエストオブジェクト
        motivation - free_description (str): 各セクションのフォームデータ
        session_id (str | None): CookieセッションID
        db (AsyncSession): データベースセッション

    Returns:
        TemplateResponse: `components/plan_viewer.html`
    """
    if not session_id:
        return HTMLResponse("Session not found", status_code=400)

    # 現在のセッションデータをロード
    data = await session_store.load_session(db, session_id)
    if not data:
        data = {"task_states": {}, "chat_history": []}

    # フォームデータからセクション辞書を構築
    # strip()は前後の空白のみ削除し、内部の改行は保持する
    sections = {
        "motivation": motivation.strip() if motivation.strip() else None,
        "background": background.strip() if background.strip() else None,
        "service": service.strip() if service.strip() else None,
        "employees": employees.strip() if employees.strip() else None,
        "partners": partners.strip() if partners.strip() else None,
        "related_companies": related_companies.strip()
        if related_companies.strip()
        else None,
        "loans": loans.strip() if loans.strip() else None,
        "funds": funds.strip() if funds.strip() else None,
        "outlook": outlook.strip() if outlook.strip() else None,
        "free_description": free_description.strip()
        if free_description.strip()
        else None,
    }

    # タスク状態をTASKS形式に変換
    task_states = data.get("task_states", {})
    current_tasks = []
    for task in TASKS:
        task_copy = task.copy()
        if task["id"] in task_states:
            task_copy["status"] = task_states[task["id"]]
        current_tasks.append(task_copy)

    # 保存
    history = data.get("chat_history", [])
    await session_store.save_session(
        db, session_id, current_tasks, history, sections=sections
    )

    # 表示用に全文テキストを再構築
    plan_text = build_plan_text_from_sections(sections)

    # 閲覧モードのHTMLを返す
    return templates.TemplateResponse(
        "components/plan_viewer.html",
        {"request": request, "plan_text": plan_text, "stepper_oob": ""},
    )


@app.post(
    "/plan/generate",
    response_class=HTMLResponse,
    summary="計画書のドラフト生成",
    description="""
    これまでのチャット内容を元に、創業計画書のドラフトを生成します。

    ⚠️ **Swagger UIでの直接テストはできません**
    - このエンドポイントはCookieによるセッション管理が必要です
    - Swagger UIではCookieパラメータを正しく送信できません

    **テスト方法:**
    - 下記の「Try it out」→「session_id」を入力→「Execute」で表示されるCurlコマンドをコピー
    - ターミナルで実行
    """,
)
async def generate_plan(
    request: Request,
    session_id: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    """
    これまでのチャット内容を元に、創業計画書のドラフトを生成します。

    Gemini APIを使用して、会話履歴全体から創業計画書形式のテキストを生成し、
    各セクションに分割してデータベースに保存します。
    また、ステッパー（進捗表示）を「創業計画書作成」フェーズに進めます。

    Args:
        request (Request): FastAPIのリクエストオブジェクト
        session_id (str | None): CookieセッションID
        db (AsyncSession): データベースセッション

    Returns:
        TemplateResponse: 生成された計画書を表示する `components/plan_viewer.html`
    """
    if not session_id:
        # セッションがない場合はエラー表示
        return HTMLResponse(
            '<div class="flex items-center justify-center h-full text-red-500">セッションが見つかりません。最初からやり直してください。</div>',
            status_code=400,
        )

    # データベースからセッションデータをロードし、チャット履歴を復元
    current_data = await session_store.load_session(db, session_id)
    if not current_data:
        return HTMLResponse(
            '<div class="flex items-center justify-center h-full text-red-500">セッションデータが見つかりません。最初からやり直してください。</div>',
            status_code=400,
        )

    # チャット履歴をGeminiServiceに復元
    history = current_data.get("chat_history", [])
    if history:
        gemini_service.restore_chat_session(session_id, history)

    # Geminiでドラフト生成
    plan_text = await gemini_service.generate_business_plan(session_id)

    # マーカーを除去（表示用に不要な制御コマンドを削除）
    plan_text = re.sub(r"\[\[DRAFT_PROPOSED\]\]", "", plan_text)
    plan_text = re.sub(r"\[\[COMPLETED:[a-z_]+\]\]", "", plan_text)

    # --- 既存のセクションデータを取得 ---
    # 注: current_data は既に上でロード済み
    existing_sections = {}
    if current_data and current_data.get("sections"):
        existing_sections = current_data["sections"]

    # 全文テキストをセクション別に分割
    new_sections = extract_sections_from_text(plan_text)

    # 既存のセクションデータと新規生成データをマージ（既存データを優先）
    sections = {}
    for key in [
        "motivation",
        "background",
        "service",
        "employees",
        "partners",
        "related_companies",
        "loans",
        "funds",
        "outlook",
        "free_description",
    ]:
        # 既存データがあればそれを使用、なければ新規生成データを使用
        if existing_sections.get(key):
            sections[key] = existing_sections[key]
        elif new_sections.get(key):
            sections[key] = new_sections[key]
        else:
            sections[key] = None

    # --- 生成されたプランの保存 ---
    if current_data:
        # タスク状態をTASKS形式に変換
        task_states = current_data.get("task_states", {})
        saved_tasks = []
        for task in TASKS:
            task_copy = task.copy()
            if task["id"] in task_states:
                task_copy["status"] = task_states[task["id"]]
            saved_tasks.append(task_copy)

        saved_history = current_data.get("chat_history", [])
        await session_store.save_session(
            db, session_id, saved_tasks, saved_history, sections=sections
        )

    # 生成した計画書をAIで自動検証（チャットセッションとは独立した単発の検証）
    verification = await gemini_service.verify_business_plan(sections)

    # ステップの状態を更新 (Conceptual: 1->Completed, 2->Current)
    current_steps = [s.copy() for s in ROADMAP_STEPS]
    current_steps[0]["status"] = "completed"
    current_steps[1]["status"] = "current"

    # ステッパーHTMLをOOB更新用にレンダリング
    stepper_html = templates.get_template("components/stepper.html").render(
        {"steps": current_steps}
    )
    # stepper.html内のnav要素をhx-swap-oob="true"にするために置換 (またはtemplate側で対応)
    # ここでは簡易的に文字列置換で hx-swap-oob 属性を注入
    stepper_html = stepper_html.replace(
        '<nav aria-label="Progress" id="stepper-nav">',
        '<nav aria-label="Progress" id="stepper-nav" hx-swap-oob="true">',
    )

    return templates.TemplateResponse(
        "components/plan_viewer.html",
        {
            "request": request,
            "plan_text": plan_text,
            "stepper_oob": stepper_html,
            "verification": verification,
        },
    )


@app.get(
    "/plan/download_excel",
    summary="計画書のExcelダウンロード",
    description="""
    作成完了した創業計画書をExcel形式でZIP圧縮してダウンロードします。

    ⚠️ **Swagger UIでの直接テストはできません**
    - このエンドポイントはCookieによるセッション管理が必要です
    - Swagger UIではCookieパラメータを正しく送信できません

    **テスト方法1: curlでファイル保存（推奨）**
    ```bash
    curl -X 'GET' \\
      'http://localhost:8000/plan/download_excel' \\
      -H 'Cookie: session_id=あなたのセッションID' \\
      -o 創業計画書.zip
    ```

    **テスト方法2: ブラウザで直接アクセス（最も簡単）**
    - ブラウザで `http://localhost:8000/plan/download_excel` を開く
    - 自動的にダウンロードが開始されます
    """,
)
async def download_plan_excel(
    session_id: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    """
    作成完了した創業計画書をExcel形式でZIP圧縮してダウンロードします。

    1. セッションから計画書の各セクションデータを取得
    2. `templates/startup_plan_template.xlsx` を読み込み、対応するセルに転記
    3. 計画書の内容から業種を推測し、適切な記入例PDF（`static/templates/examples/*.pdf`）を選択
    4. ExcelファイルとPDFファイルをZIPアーカイブにまとめてストリーミング返却

    Args:
        session_id (str | None): CookieセッションID
        db (AsyncSession): データベースセッション

    Returns:
        StreamingResponse: ZIPファイル（application/zip）
    """
    if not session_id:
        return Response("Session not found", status_code=400)

    data = await session_store.load_session(db, session_id)
    if not data or not data.get("sections"):
        return Response("Plan data not found", status_code=404)

    sections = data.get("sections")

    # テンプレート読み込み
    template_path = BASE_DIR / "static" / "templates" / "startup_plan_template.xlsx"
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active

    # 作成日を令和形式で書き込み（AC1=年, AF1=月, AI1=日）
    _today = date.today()
    _reiwa_year = _today.year - 2018  # 令和元年 = 2019年
    sheet["AC1"].value = _reiwa_year
    sheet["AF1"].value = _today.month
    sheet["AI1"].value = _today.day

    # テンプレート内の見出しセルを探して、転記先セルを動的に推定する
    label_to_key = {
        "創業の動機": "motivation",
        "経営者の略歴等": "background",
        "取扱商品・サービス": "service",
        "従業員": "employees",
        "取引先・取引関係等": "partners",
        "関連企業": "related_companies",
        "お借入の状況": "loans",
        "必要な資金と調達方法": "funds",
        "事業の見通し": "outlook",
        "自由記述欄": "free_description",
    }

    label_cells = {}
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                value = cell.value.strip()
                for label, key in label_to_key.items():
                    if label in value:
                        label_cells[key] = cell

    def choose_target_cell(label_cell):
        label_row = label_cell.row
        label_col = label_cell.column
        candidates = []
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            if min_row >= label_row + 1 and min_row <= label_row + 12:
                width = max_col - min_col + 1
                height = max_row - min_row + 1
                if width >= 6:
                    if min_col <= label_col <= max_col or min_col > label_col:
                        candidates.append(
                            (min_row, -(width * height), min_col, str(merged_range))
                        )
        if candidates:
            candidates.sort()
            selected = candidates[0][3]
            min_col, min_row, _, _ = range_boundaries(selected)
            return f"{get_column_letter(min_col)}{min_row}"
        return f"{get_column_letter(label_col + 1)}{label_row}"

    mapping = {key: None for key in label_to_key.values()}
    for key, cell in label_cells.items():
        mapping[key] = choose_target_cell(cell)

    # フォールバック用の静的マッピング（テンプレート検出できない場合）
    fallback_mapping = {
        "motivation": "A9",
        "background": "A16",
        "service": "A27",
        "employees": "T21",
        "partners": "M27",
        "related_companies": "A36",
        "loans": "M36",
        "funds": "A45",
        "outlook": "M45",
        "free_description": "A54",
    }
    for key, addr in fallback_mapping.items():
        if not mapping.get(key):
            mapping[key] = addr

    # 特定のセルを手動で上書き（動的検出が正しくない場合）
    mapping["background"] = "H13"  # 経営者の略歴等
    mapping["employees"] = "W43"  # 従業員数 入力セル（フォールバック用）
    mapping["partners"] = "C49"  # 取引先名 第1入力行（フォールバック用）
    mapping["funds"] = "AP20"  # 設備資金内訳 第1テキストセル（フォールバック用）
    # mapping["outlook"] は数値抽出処理のみ行うため、フォールバック用のマッピングは削除

    # データベースから取得したセクションデータをExcelに転記
    for key, content in sections.items():
        if content:
            # 「創業の動機」は特殊処理（60文字ごとに分割してB7～B12に転記）
            if key == "motivation":
                # 改行を削除して1つの文字列にする
                text = content.replace("\n", "")

                # 60文字ごとに分割（最大6行）
                max_rows = 6
                char_limit = 60

                for idx in range(max_rows):
                    start_pos = idx * char_limit
                    end_pos = start_pos + char_limit

                    if start_pos >= len(text):
                        break

                    line_text = text[start_pos:end_pos]
                    row_num = 7 + idx  # B7, B8, B9, B10, B11, B12

                    sheet[f"B{row_num}"].value = line_text

                continue

            # 「経営者の略歴等」は特殊処理（略歴と資格・許認可欄を分けて転記）
            if key == "background":
                lines = content.strip().split("\n")

                # 略歴行と特殊フィールドに振り分け
                history_lines = []
                qualification_text = ""  # 取得資格
                permit_text = ""  # 許認可
                ip_text = ""  # 知的財産権等

                for line in lines:
                    line_s = line.strip()
                    if line_s.startswith("取得資格："):
                        qualification_text = line_s[len("取得資格：") :]
                    elif line_s.startswith("許認可："):
                        permit_text = line_s[len("許認可：") :]
                    elif line_s.startswith("知的財産権等："):
                        ip_text = line_s[len("知的財産権等：") :]
                    elif line_s.startswith("過去の事業経験："):
                        pass  # チェックボックス形式のため書き込み対象外
                    elif line_s:
                        history_lines.append(line_s)

                # 略歴行（最大4行: rows 13-16, B: 年月, H: 内容）
                for idx, line in enumerate(history_lines[:4]):
                    line = line.lstrip("・-− ")
                    row_num = 13 + idx

                    year_month_match = re.match(
                        r"^([0-9]{4}年[0-9]{1,2}月|[平成令和昭和]+[0-9元]{1,2}年[0-9]{1,2}月)",
                        line,
                    )

                    if year_month_match:
                        year_month = year_month_match.group(1)
                        # 年月の直後の「：」「:」区切り文字を除去する
                        content_text = (
                            line[len(year_month) :].strip().lstrip("：:").strip()
                        )
                    else:
                        year_month = ""
                        content_text = line

                    if year_month:
                        sheet[f"B{row_num}"].value = year_month
                    if content_text:
                        sheet[f"H{row_num}"].value = content_text

                if qualification_text:
                    sheet["P22"].value = qualification_text
                if permit_text:
                    sheet["P23"].value = permit_text
                if ip_text:
                    sheet["P24"].value = ip_text

                continue

            # 「取扱商品・サービス」は複数セルに振り分けて転記
            if key == "service":
                # セクション境界キーワード: (section_name, pattern)
                _SERVICE_SECTIONS = [
                    ("products", r"取扱商品.*サービス.*内容"),
                    ("sales_point", r"セールスポイント|自社の強み"),
                    ("target", r"販売ターゲット|販売戦略|集客方法"),
                    ("market", r"競合.{0,5}市場|自社を取り巻く"),
                ]

                current_section = "description"
                description_lines = []
                products = []  # (説明文, 売上シェア%)
                sales_point_lines = []
                target_lines = []
                market_lines = []

                for raw_line in content.strip().split("\n"):
                    line = raw_line.strip()
                    if not line:
                        continue

                    # セクション遷移キーワードを検出
                    # ヘッダーの後ろに内容が続く場合（"：内容"）も取り出す
                    matched_section = None
                    after_header = ""
                    for section_name, pattern in _SERVICE_SECTIONS:
                        m = re.search(pattern, line)
                        if m:
                            matched_section = section_name
                            # "：" or ":" 以降の文字列を取り出す
                            colon_m = re.search(r"[：:]\s*(.+)$", line[m.end() :])
                            if colon_m:
                                after_header = colon_m.group(1).strip()
                            break

                    if matched_section:
                        current_section = matched_section
                        if not after_header:
                            continue
                        # ヘッダー同行に内容がある場合はその内容で処理続行
                        line = after_header

                    # セクションに応じて振り分け
                    if current_section == "description":
                        line_clean = re.sub(r"^事業内容[：:]\s*", "", line)
                        if line_clean:
                            description_lines.append(line_clean)
                    elif current_section == "products":
                        m = re.match(r"^[①②③１２３123]\s*(.*)", line)
                        if m:
                            item_text = m.group(1).strip()
                            share_m = re.search(
                                r"[（(]売上シェア\s*(\d+)\s*[%％][）)]", item_text
                            )
                            if share_m:
                                share_pct = share_m.group(1)
                                item_text = item_text[: share_m.start()].strip()
                            else:
                                share_pct = ""
                            products.append((item_text, share_pct))
                        elif re.match(r"^客単価[：:]", line):
                            val = re.sub(r"^客単価[：:]\s*", "", line)
                            sheet["H31"].value = val
                    elif current_section == "sales_point":
                        sales_point_lines.append(line)
                    elif current_section == "target":
                        target_lines.append(line)
                    elif current_section == "market":
                        market_lines.append(line)

                # 事業内容 → H26, H27（50文字ごとに折り返し）
                def _wrap50(lines, width=50):
                    result = []
                    for part in lines:
                        while len(part) > width:
                            result.append(part[:width])
                            part = part[width:]
                        if part:
                            result.append(part)
                    return result

                wrapped_desc = _wrap50(description_lines)
                if wrapped_desc:
                    sheet["H26"].value = wrapped_desc[0]
                    if len(wrapped_desc) > 1:
                        sheet["H27"].value = "\n".join(wrapped_desc[1:3])

                # ①②③ → I28/I29/I30、シェア% → AJ28/AJ29/AJ30
                for idx, (item_text, share_pct) in enumerate(products[:3]):
                    row_num = 28 + idx
                    if item_text:
                        sheet[f"I{row_num}"].value = item_text
                    if share_pct:
                        sheet[f"AJ{row_num}"].value = share_pct

                # セールスポイント → H33, H34, H35（50文字折り返し・超過分は最終行に結合）
                wrapped_sales = _wrap50(sales_point_lines)
                for idx in range(3):
                    if idx < len(wrapped_sales):
                        val = (
                            "\n".join(wrapped_sales[idx:])
                            if idx == 2
                            else wrapped_sales[idx]
                        )
                        sheet[f"H{33 + idx}"].value = val

                # 販売ターゲット・販売戦略 → H36, H37, H38（50文字折り返し）
                wrapped_target = _wrap50(target_lines)
                for idx in range(3):
                    if idx < len(wrapped_target):
                        val = (
                            "\n".join(wrapped_target[idx:])
                            if idx == 2
                            else wrapped_target[idx]
                        )
                        sheet[f"H{36 + idx}"].value = val

                # 競合・市場 → H39, H40, H41（50文字折り返し）
                wrapped_market = _wrap50(market_lines)
                for idx in range(3):
                    if idx < len(wrapped_market):
                        val = (
                            "\n".join(wrapped_market[idx:])
                            if idx == 2
                            else wrapped_market[idx]
                        )
                        sheet[f"H{39 + idx}"].value = val

                continue

            # 「4. 従業員」は数値パースで各セルに書き込み
            if key == "employees":
                # 1行ずつ処理することでパターンの競合（家族従業員がW43に入るなど）を防ぐ
                # 「常勤役員の人数（法人のみ）：1名」のように修飾語が入っても対応できる
                _emp_line_patterns = [
                    (r"常勤役員[^：:\n]*[：:]\s*(\d+)名?", "I43"),
                    (r"(?<!家族)(?:従業員[数]?|社員)[^：:\n]*[：:]\s*(\d+)名?", "W43"),
                    (r"(?:家族従業員|(?:うち)?家族)[^：:\n]*[：:]\s*(\d+)名?", "AH43"),
                    (r"パート[^：:\n]*[：:]\s*(\d+)名?", "AH44"),
                ]
                for _emp_line in content.split("\n"):
                    _emp_line = _emp_line.strip()
                    if not _emp_line:
                        continue
                    for _pat, _cell in _emp_line_patterns:
                        _m = re.search(_pat, _emp_line)
                        if _m:
                            _num = int(_m.group(1))
                            # 0人の場合は空欄のままにする（テンプレートの「人」ラベルと重複しない）
                            if _num > 0:
                                try:
                                    sheet[_cell].value = _num
                                except Exception:
                                    pass
                continue

            # 「5. 取引先・取引関係等」は販売先/仕入先/外注先ごとに正しいセルへ書き込み
            if key == "partners":
                # セクション別セルマッピング（フリガナ行, 取引先名行, 所在地, シェア, 掛取引割合, 締日, 支払日）
                # 注: すべてマージセルの左上セルを指定
                # X列=掛取引の割合, AD列=締日（「末」「15」等）, AI列=支払日（「翌月10」等）
                _PARTNER_CELLS = {
                    "販売先": [
                        {
                            "kana": "C49",
                            "name": "C50",
                            "location": "O49",
                            "share": "U49",
                            "credit_rate": "X49",
                            "closing": "AD49",
                            "payment": "AI49",
                        },
                        {
                            "kana": "C51",
                            "name": "C52",
                            "location": "O51",
                            "share": "U51",
                            "credit_rate": "X51",
                            "closing": "AD51",
                            "payment": "AI51",
                        },
                    ],
                    "仕入先": [
                        {
                            "kana": "C55",
                            "name": "C56",
                            "location": "O55",
                            "share": "U55",
                            "credit_rate": "X55",
                            "closing": "AD55",
                            "payment": "AI55",
                        },
                        {
                            "kana": "C57",
                            "name": "C58",
                            "location": "O57",
                            "share": "U57",
                            "credit_rate": "X57",
                            "closing": "AD57",
                            "payment": "AI57",
                        },
                    ],
                    "外注先": [
                        {
                            "kana": "C61",
                            "name": "C62",
                            "location": "O61",
                            "share": "U61",
                            "credit_rate": "X61",
                            "closing": "AD61",
                            "payment": "AI61",
                        },
                    ],
                }
                _P_NASHI = {
                    "なし",
                    "特になし",
                    "該当なし",
                    "ありません",
                    "ない",
                    "なし。",
                    "特になし。",
                    "該当なし。",
                }
                _sec_counts = {k: 0 for k in _PARTNER_CELLS}
                _current_sec = None

                for _raw in content.split("\n"):
                    _pline = _raw.strip()
                    if not _pline:
                        continue

                    # セクション見出し検出（「販売先：」「仕入先」「外注先」）
                    for _sec in ["販売先", "仕入先", "外注先"]:
                        if re.match(rf"^{_sec}", _pline):
                            _current_sec = _sec
                            # 見出し後に同行データがある場合（「販売先：〇〇社」）
                            _pline = re.sub(rf"^{_sec}[：:\s]*", "", _pline).strip()
                            break

                    if not _pline or _current_sec is None:
                        continue

                    # 行頭の記号（・―-）を先に除去してからなし判定
                    _pline = re.sub(r"^[・\-\−●○]+\s*", "", _pline)
                    if not _pline:
                        continue

                    # 「なし」系はスキップ
                    if _pline.replace("　", "").replace(" ", "") in _P_NASHI:
                        continue

                    # シェア・支払条件のみの行は会社名として登録しない
                    if re.match(
                        r"^(?:シェア|取引シェア|掛取引|支払条件|末締め|月末締め|即時払い|前払い|翌月払い)",
                        _pline,
                    ):
                        continue

                    # 最大登録数を超えたらスキップ
                    _pidx = _sec_counts.get(_current_sec, 0)
                    _pcells_list = _PARTNER_CELLS.get(_current_sec, [])
                    if _pidx >= len(_pcells_list):
                        continue

                    _pcells = _pcells_list[_pidx]

                    # 所在地の抽出（括弧付き or スペース区切り、日本・海外問わず）
                    _ploc = ""
                    _pcomp = _pline

                    # 1. 括弧付きパターン（「（東京都）」「（首都ハボロネ）」など）
                    _loc_m = re.search(r"（([^）]+)）", _pline)
                    if _loc_m:
                        _ploc = _loc_m.group(1).strip()
                        _pcomp = _pline[: _loc_m.start()].strip()
                    else:
                        # 2. スペース区切りパターン（「クラスメソッド 東京都港区」「個人 首都ハボロネ」等）
                        # シェア/カンマ等のキーワード前までを対象にする
                        _before_kw = re.split(
                            r"(?:\s+シェア|\s+取引シェア|[、,，])", _pline
                        )[0]
                        _parts = _before_kw.split(None, 1)  # 最初の空白で2分割

                        _pcomp = _parts[0] if _parts else _pline
                        if len(_parts) >= 2:
                            # 2番目の部分から所在地を取得（残りのキーワードを除外）
                            _loc_text = _parts[1].strip()
                            _ploc = re.split(
                                r"\s+(?:シェア|掛取引|支払条件|即金|即時払い|前払い|翌月払い|末締め|月末締め)",
                                _loc_text,
                            )[0].strip()

                    # シェア%パース
                    _pshare_m = re.search(
                        r"(?:シェア|取引シェア|取引先のシェア)[：:\s]*(\d+)\s*[%％]",
                        _pline,
                    )

                    # 掛取引の割合パース
                    # 様々な表記に対応: 「掛取引100%」「掛け取引の割合50％」「掛100%」
                    _credit_rate = None
                    _credit_m = re.search(
                        r"(?:掛け?取引(?:の割合)?|掛け?)[：:\s]*(\d+)\s*[%％]", _pline
                    )
                    if _credit_m:
                        _credit_rate = int(_credit_m.group(1))
                    # 即金系キーワード → 掛取引0%
                    elif re.search(
                        r"(?:即金|現金(?:取引)?|即時払い|即日|前払い|即払い)", _pline
                    ):
                        _credit_rate = 0

                    # 回収・支払条件のパース
                    # AD列：締日（「末」「15」など）
                    # AI列：支払日（「翌月10」「翌30」など）
                    _closing_text = ""
                    _payment_text = ""

                    # 即金パターンをまず確認
                    if re.search(r"即金|即時払い", _pline):
                        _closing_text = "即金"
                    else:
                        # 締日の抽出（「末日締め」→「末」、「15日締め」→「15」）
                        _closing_m = re.search(r"(月末|末日|末|(\d{1,2})日?)締", _pline)
                        if _closing_m:
                            if _closing_m.group(2):  # 数字がある場合（15日締め等）
                                _closing_text = _closing_m.group(2)
                            else:  # 「月末」「末日」「末」
                                _closing_text = "末"

                        # 支払日の抽出（「翌月10日払い」→「翌月10」、「翌30日回収」→「翌30」）
                        _payment_day_m = re.search(
                            r"(翌々?月|翌|当月)?(\d{1,2})日(?:払い|支払|回収)", _pline
                        )
                        if _payment_day_m:
                            _prefix = (
                                _payment_day_m.group(1)
                                if _payment_day_m.group(1)
                                else ""
                            )
                            _day = _payment_day_m.group(2)
                            _payment_text = f"{_prefix}{_day}"
                        # 月末払いパターン
                        elif re.search(
                            r"(翌々?月|当月)?月末(?:払い|支払|回収)", _pline
                        ):
                            _month_m = re.search(r"(翌々?月|当月)?月末", _pline)
                            _prefix = _month_m.group(1) if _month_m.group(1) else ""
                            _payment_text = f"{_prefix}月末"

                    # セルへ書き込み
                    sheet[_pcells["name"]].value = _pcomp
                    if _ploc:
                        sheet[_pcells["location"]].value = _ploc
                    if _pshare_m:
                        try:
                            sheet[_pcells["share"]].value = int(_pshare_m.group(1))
                        except Exception:
                            pass
                    if _credit_rate is not None:
                        try:
                            sheet[_pcells["credit_rate"]].value = _credit_rate
                        except Exception:
                            pass
                    if _closing_text:
                        try:
                            sheet[_pcells["closing"]].value = _closing_text
                        except Exception:
                            pass
                    if _payment_text:
                        try:
                            sheet[_pcells["payment"]].value = _payment_text
                        except Exception:
                            pass

                    _sec_counts[_current_sec] = _pidx + 1

                # 人件費の支払情報を抽出（G65とM65に書き込む）
                for _raw in content.split("\n"):
                    _pline = _raw.strip()
                    if re.match(r"人件費の?支払", _pline):
                        # 締日の抽出
                        _personnel_closing = ""
                        _personnel_payment = ""

                        # 即金パターン
                        if re.search(r"即金|即時払い", _pline):
                            _personnel_closing = "即金"
                        else:
                            # 締日の抽出（「末日締め」→「末」、「15日締め」→「15」）
                            _closing_m = re.search(
                                r"(月末|末日|末|(\d{1,2})日?)締", _pline
                            )
                            if _closing_m:
                                if _closing_m.group(2):  # 数字がある場合
                                    _personnel_closing = _closing_m.group(2)
                                else:  # 「月末」「末日」「末」
                                    _personnel_closing = "末"

                            # 支払日の抽出（「翌月10日払い」→「翌月10」）
                            _payment_day_m = re.search(
                                r"(翌々?月|翌|当月)?(\d{1,2})日(?:払い|支払)", _pline
                            )
                            if _payment_day_m:
                                _prefix = (
                                    _payment_day_m.group(1)
                                    if _payment_day_m.group(1)
                                    else ""
                                )
                                _day = _payment_day_m.group(2)
                                _personnel_payment = f"{_prefix}{_day}"
                            # 月末払いパターン
                            elif re.search(r"(翌々?月|当月)?月末(?:払い|支払)", _pline):
                                _month_m = re.search(r"(翌々?月|当月)?月末", _pline)
                                _prefix = _month_m.group(1) if _month_m.group(1) else ""
                                _personnel_payment = f"{_prefix}月末"

                        # G65とM65に書き込み
                        if _personnel_closing:
                            try:
                                sheet["G65"].value = _personnel_closing
                            except Exception:
                                pass
                        if _personnel_payment:
                            try:
                                sheet["M65"].value = _personnel_payment
                            except Exception:
                                pass
                        break  # 人件費の支払は1行のみなので抜ける

                continue

            # 「6. 関連企業」が「なし」系の回答の場合は空欄のままにする
            if key == "related_companies":
                # 「関連企業：」などのプレフィックスを除去してから判定
                _rc_text = content.strip()
                _rc_text = re.sub(r"^関連企業[：:\s]*", "", _rc_text)
                _rc_normalized = (
                    _rc_text.replace("　", "").replace(" ", "").replace("。", "")
                )
                _NASHI_SET = {"なし", "特になし", "該当なし", "ありません", "ない"}
                if _rc_normalized in _NASHI_SET:
                    continue  # 書き込まずにスキップ
                # 通常通り書き込み
                _rc_cell = mapping.get(key)
                if _rc_cell:
                    try:
                        sheet[_rc_cell].value = content
                    except Exception as e:
                        print(f"[ERROR] ✗ Could not write {key}: {e}")
                continue

            # 「8. 必要な資金と調達方法」は設備資金・運転資金の内訳セルに書き込み
            if key == "funds":
                # 内訳行を抽出（「・」「-」などの行頭記号がある行、または「内訳」以降の行）
                _in_breakdown = False
                _breakdown_items = []
                for _raw_line in content.split("\n"):
                    _line = _raw_line.strip()
                    if not _line:
                        continue
                    # テンプレートに既に入っている見出し行をスキップ
                    if re.match(r"^(?:必要な資金|設備資金|運転資金)[：:]", _line):
                        continue
                    # 「商品仕入、経費支払資金など」などのラベルをスキップ
                    if re.match(
                        r"^(?:店舗、工場、機械、車両など|商品仕入、経費支払資金など)",
                        _line,
                    ):
                        continue
                    # 「内訳」「（内訳）」で内訳開始
                    if re.match(r"^[（(]?内訳[）)]?$", _line):
                        _in_breakdown = True
                        continue
                    # 内訳モード中、または行頭記号がある行、または括弧で始まる補足説明を内訳とみなす
                    if _in_breakdown or re.match(r"^[・\-\−●○（(]\s*", _line):
                        _breakdown_items.append(_line)

                # 設備資金内訳 (AP20-AP22) と運転資金内訳 (AP32-AP34)
                # 項目名=AP列、金額=BD列
                _equipment_cells = [
                    ("AP20", "BD20"),
                    ("AP21", "BD21"),
                    ("AP22", "BD22"),
                ]
                _working_cells = [("AP32", "BD32"), ("AP33", "BD33"), ("AP34", "BD34")]

                # シンプルに最初の3項目を設備資金、次の3項目を運転資金とする
                _equipment_items = _breakdown_items[:3]
                _working_items = _breakdown_items[3:6]

                # 設備資金の処理
                for i, _item_line in enumerate(_equipment_items):
                    if i >= len(_equipment_cells):
                        break

                    _item_line = re.sub(r"^[・\-\−●○]\s*", "", _item_line)
                    _amount_m = re.search(
                        r"(\d+(?:,\d{3})*)\s*(?:万\s*円|円|万)", _item_line
                    )
                    _item_name = _item_line
                    _amount_num = None

                    if _amount_m:
                        _amount_str = _amount_m.group(1).replace(",", "")
                        try:
                            _amount_num = int(_amount_str)
                        except ValueError:
                            pass
                        _item_name = _item_line[: _amount_m.start()].strip()

                    _item_cell, _amount_cell = _equipment_cells[i]
                    try:
                        sheet[_item_cell].value = _item_name
                    except Exception:
                        pass
                    if _amount_num is not None:
                        try:
                            sheet[_amount_cell].value = _amount_num
                        except Exception:
                            pass

                # 運転資金の処理
                for i, _item_line in enumerate(_working_items):
                    if i >= len(_working_cells):
                        break

                    _item_line = re.sub(r"^[・\-\−●○]\s*", "", _item_line)
                    _amount_m = re.search(
                        r"(\d+(?:,\d{3})*)\s*(?:万\s*円|円|万)", _item_line
                    )
                    _item_name = _item_line
                    _amount_num = None

                    if _amount_m:
                        _amount_str = _amount_m.group(1).replace(",", "")
                        try:
                            _amount_num = int(_amount_str)
                        except ValueError:
                            pass
                        _item_name = _item_line[: _amount_m.start()].strip()

                    _item_cell, _amount_cell = _working_cells[i]
                    try:
                        sheet[_item_cell].value = _item_name
                    except Exception:
                        pass
                    if _amount_num is not None:
                        try:
                            sheet[_amount_cell].value = _amount_num
                        except Exception:
                            pass

                # 自己資金・公庫借入額のパース（カンマ区切り数字に対応）
                _sf_m = re.search(r"自己資金[：:\s]*(\d+(?:,\d{3})*)万?円?", content)
                if _sf_m:
                    try:
                        _sf_amount = int(_sf_m.group(1).replace(",", ""))
                        sheet["BV18"].value = _sf_amount
                    except Exception:
                        pass
                # 「日本政策金融公庫 国民生活事業からの借入 1,000万円」のような形式に対応
                _jfc_m = re.search(
                    r"(?:公庫|日本政策).*?(?:借入|融資).*?(\d+(?:,\d{3})*)万?円?",
                    content,
                )
                if _jfc_m:
                    try:
                        _jfc_amount = int(_jfc_m.group(1).replace(",", ""))
                        sheet["BV24"].value = _jfc_amount
                    except Exception:
                        pass
                continue

            # 「9. 事業の見通し」は4ブロックに分けて処理
            if key == "outlook":
                # ブロック1: 「創業当初：」から「１年後又は軌道に乗った後：」の前まで
                # ブロック2: 「１年後又は軌道に乗った後：」から「＜創業当初＞」の前まで
                # ブロック3: 「＜創業当初＞」から「＜軌道に乗った後＞」の前まで
                # ブロック4: 「＜軌道に乗った後＞」から最後まで

                # ブロック1: 創業当初の数値部分を抽出
                _block1_match = re.search(
                    r"創業当初：(.*?)(?=１年後又は軌道に乗った後：|＜創業当初＞|$)",
                    content,
                    re.DOTALL
                )
                _block1 = _block1_match.group(1) if _block1_match else ""

                # ブロック2: 軌道に乗った後の数値部分を抽出
                _block2_match = re.search(
                    r"１年後又は軌道に乗った後：(.*?)(?=＜創業当初＞|＜軌道に乗った後＞|$)",
                    content,
                    re.DOTALL
                )
                _block2 = _block2_match.group(1) if _block2_match else ""

                # ブロック3: 創業当初の説明部分を抽出
                _block3_match = re.search(
                    r"＜創業当初＞(.*?)(?=＜軌道に乗った後＞|$)",
                    content,
                    re.DOTALL
                )
                _block3 = _block3_match.group(1).strip() if _block3_match else ""

                # ブロック4: 軌道に乗った後の説明部分を抽出
                _block4_match = re.search(
                    r"＜軌道に乗った後＞(.*?)$",
                    content,
                    re.DOTALL
                )
                _block4 = _block4_match.group(1).strip() if _block4_match else ""

                # ブロック1から創業当初の数値を抽出してAT列に書き込み
                _initial_patterns = [
                    (r"売上高[^0-9\n]*(\d+(?:,\d{3})*)", "AT41"),
                    (r"売上原価[^0-9\n]*(\d+(?:,\d{3})*)", "AT44"),
                    (r"人件費[^0-9\n]*(\d+(?:,\d{3})*)", "AT46"),
                    (r"家賃[^0-9\n]*(\d+(?:,\d{3})*)", "AT48"),
                    (r"支払利息[^0-9\n]*(\d+(?:,\d{3})*)", "AT50"),
                    (r"その他[^0-9\n]*(\d+(?:,\d{3})*)", "AT52"),
                ]
                for _pattern, _cell in _initial_patterns:
                    _m = re.search(_pattern, _block1)
                    if _m:
                        try:
                            sheet[_cell].value = int(_m.group(1).replace(",", ""))
                        except Exception:
                            pass

                # ブロック2から軌道に乗った後の数値を抽出してAZ列に書き込み
                _later_patterns = [
                    (r"売上高[^0-9\n]*(\d+(?:,\d{3})*)", "AZ41"),
                    (r"売上原価[^0-9\n]*(\d+(?:,\d{3})*)", "AZ44"),
                    (r"人件費[^0-9\n]*(\d+(?:,\d{3})*)", "AZ46"),
                    (r"家賃[^0-9\n]*(\d+(?:,\d{3})*)", "AZ48"),
                    (r"支払利息[^0-9\n]*(\d+(?:,\d{3})*)", "AZ50"),
                    (r"その他[^0-9\n]*(\d+(?:,\d{3})*)", "AZ52"),
                ]
                for _pattern, _cell in _later_patterns:
                    _m = re.search(_pattern, _block2)
                    if _m:
                        try:
                            sheet[_cell].value = int(_m.group(1).replace(",", ""))
                        except Exception:
                            pass

                # ブロック3とブロック4をBF41に書き込み
                _explanation_text = ""
                if _block3:
                    _explanation_text += "＜創業当初＞\n" + _block3
                if _block4:
                    if _explanation_text:
                        _explanation_text += "\n"
                    _explanation_text += "＜軌道に乗った後＞\n" + _block4

                if _explanation_text:
                    sheet["BF41"].value = _explanation_text.strip()

                continue

            # その他のセクションは通常通り処理
            cell_addr = mapping.get(key)
            if cell_addr:
                try:
                    sheet[cell_addr].value = content
                except Exception as e:
                    print(f"[ERROR] ✗ Could not write {key} to cell {cell_addr}: {e}")
            else:
                print(f"[WARNING] No cell mapping found for {key}")

    # メモリ上のバイナリとして保存 (Excel)
    excel_output = BytesIO()
    workbook.save(excel_output)
    excel_output.seek(0)

    # 業種判定とPDF選択
    examples_dir = BASE_DIR / "static" / "templates" / "examples"
    pdf_filename = None

    # 業種タイプ→PDFファイルのマッピング
    industry_pdf_map = {
        "software": "software_example.pdf",
        "restaurant": "restaurant_example.pdf",
        "beauty": "beauty_example.pdf",
        "apparel": "apparel_example.pdf",
        "construction": "construction_example.pdf",
        "cram_school": "cram_school_example.pdf",
        "care_service": "care_service_example.pdf",
        "car_sales": "car_sales_example.pdf",
        "dentist": "dentist_example.pdf",
    }

    # セッションに保存された業種タイプを取得
    industry_type = data.get("industry_type")
    if industry_type and industry_type in industry_pdf_map:
        pdf_filename = industry_pdf_map[industry_type]
    else:
        # 業種タイプが保存されていない場合はエラー
        print(
            "[WARNING] Industry type not found in session. User may have skipped selection."
        )
        # デフォルトのPDFを使用しない（業種選択は必須）
        pdf_filename = None

    # ZIPファイルの作成
    zip_output = BytesIO()
    with zipfile.ZipFile(zip_output, "w", zipfile.ZIP_DEFLATED) as zf:
        # Add Excel
        excel_filename = f"創業計画書_{session_id[:8]}.xlsx"
        zf.writestr(excel_filename, excel_output.getvalue())

        # Add PDF if found
        if pdf_filename:
            pdf_path = examples_dir / pdf_filename
            if pdf_path.exists():
                zf.write(pdf_path, arcname="創業計画書記入例.pdf")
            else:
                print(f"[WARNING] PDF file not found: {pdf_path}")
    zip_output.seek(0)
    zip_filename = f"創業計画書一式_{session_id[:8]}.zip"
    # URLエンコードしたファイル名をRFC 5987形式で指定
    from urllib.parse import quote

    encoded_filename = quote(zip_filename)
    headers = {
        "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
    }

    return StreamingResponse(
        zip_output,
        headers=headers,
        media_type="application/zip",
    )


@app.get(
    "/interview-prep",
    response_class=HTMLResponse,
    summary="面談対策画面の表示",
    description="""
    創業計画書作成後の面談対策画面を表示します。
    AI検証フィードバックと日本政策金融公庫ビジネスサポートプラザの案内を表示します。

    ⚠️ **Swagger UIでの直接テストはできません**
    - このエンドポイントはCookieによるセッション管理が必要です
    """,
)
async def interview_prep(
    request: Request,
    session_id: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    """
    面談対策画面を表示します。

    セッションに保存されている創業計画書を取得し、AI検証を実行して
    フィードバックとビジネスサポートプラザの案内を表示します。

    Args:
        request (Request): FastAPIのリクエストオブジェクト
        session_id (str | None): CookieセッションID
        db (AsyncSession): データベースセッション

    Returns:
        TemplateResponse: 面談対策画面 `components/interview_prep.html`
    """
    if not session_id:
        return HTMLResponse(
            '<div class="flex items-center justify-center h-full text-red-500">セッションが見つかりません。</div>',
            status_code=400,
        )

    # セッションデータを取得
    current_data = await session_store.load_session(db, session_id)
    if not current_data or not current_data.get("sections"):
        return HTMLResponse(
            '<div class="flex items-center justify-center h-full text-red-500">創業計画書が見つかりません。先に計画書を作成してください。</div>',
            status_code=400,
        )

    sections = current_data["sections"]

    # AI検証を実行（最新の状態で検証）
    verification = await gemini_service.verify_business_plan(sections)

    # ステッパーの状態を更新（面談対策フェーズに進める）
    current_steps = [s.copy() for s in ROADMAP_STEPS]
    current_steps[0]["status"] = "completed"  # 構想・準備
    current_steps[1]["status"] = "completed"  # 創業計画書作成
    current_steps[2]["status"] = "current"    # 面談対策

    # ステッパーHTMLをOOB更新用にレンダリング
    stepper_html = templates.get_template("components/stepper.html").render(
        {"steps": current_steps}
    )
    stepper_html = stepper_html.replace(
        '<nav aria-label="Progress" id="stepper-nav">',
        '<nav aria-label="Progress" id="stepper-nav" hx-swap-oob="true">',
    )

    return templates.TemplateResponse(
        "components/interview_prep.html",
        {
            "request": request,
            "verification": verification,
            "stepper_oob": stepper_html,
        },
    )


@app.post(
    "/reset",
    response_class=HTMLResponse,
    summary="セッションのリセット",
    description="""
    現在のセッションを完全にリセットし、初期状態に戻します。

    ⚠️ **Swagger UIでの直接テストはできません**
    - このエンドポイントはCookieによるセッション管理が必要です
    - Swagger UIではCookieパラメータを正しく送信できません

    **テスト方法:**
    - 下記の「Try it out」→「session_id」を入力→「Execute」で表示されるCurlコマンドをコピー
    - ターミナルで実行
    """,
)
async def reset_session(
    request: Request,
    session_id: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    """
    現在のセッションを完全にリセットし、初期状態に戻します。

    サーバー側のセッションデータと会話履歴を削除し、
    トップページへリダイレクトします。

    Args:
        request (Request): FastAPIのリクエストオブジェクト
        session_id (str | None): CookieセッションID
        db (AsyncSession): データベースセッション

    Returns:
        Response: HX-Redirectヘッダーを含むレスポンス
    """
    if session_id:
        # セッションデータを削除
        await session_store.delete_session(db, session_id)
        # チャットセッションもリセット
        gemini_service.reset_chat_session(session_id)

    # htmxのリダイレクト機能を使用
    from fastapi.responses import Response

    response = Response(status_code=200)
    response.headers["HX-Redirect"] = "/"
    return response

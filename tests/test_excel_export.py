"""
Test for Excel export functionality
創業計画書のExcelエクスポート機能のテスト
"""

import re
import tempfile
from pathlib import Path

import openpyxl
import pytest
from openpyxl.utils import get_column_letter, range_boundaries


class TestExcelExport:
    """創業計画書のExcelエクスポート機能をテストするクラス"""

    @pytest.fixture
    def template_path(self):
        """Excelテンプレートのパスを返す"""
        base_dir = Path(__file__).resolve().parent.parent
        return base_dir / "app" / "static" / "templates" / "startup_plan_template.xlsx"

    @pytest.fixture
    def sample_plan_text(self):
        """テスト用の事業計画書テキストを返す"""
        return """
1. 創業の動機

大学卒業後、IT企業でシステムエンジニアとして10年間勤務し、業務システムの開発に従事してきました。顧客企業の課題解決に直接貢献できることにやりがいを感じる一方で、より柔軟で迅速な開発体制を実現したいと考えるようになりました。近年、中小企業のDX推進ニーズが高まる中、大手ベンダーでは対応しきれない小規模案件が多数存在することを知り、自社で受託開発事業を立ち上げることを決意しました。

2. 経営者の略歴等

・平成XX年3月 〇〇大学工学部 卒業
・平成XX年4月 株式会社△△入社（ITシステム開発）
  - プログラマーとして業務システム開発に従事。
  - その後、システムエンジニア、プロジェクトリーダーを経験。
・令和XX年XX月 同社退職（現在に至る）
"""

    def build_section_pattern(self, current_label, next_labels):
        """セクション抽出用の正規表現パターンを構築"""
        next_part = "|".join(re.escape(label) for label in next_labels)
        return rf"(?:^|\n)\s*(?:\d+[\.|\s]*)?{re.escape(current_label)}\s*(?:\n|:|：)\s*(.*?)(?=\n\s*(?:\d+[\.|\s]*)?(?:{next_part})\s*(?:\n|:|：)|\Z)"

    def normalize_section_text(self, content: str) -> str:
        """セクションテキストを正規化"""
        content = content.strip()
        content = re.sub(r"[\*#]+", "", content)
        content = re.sub(r"\n\s*\n+", "\n\n", content)
        return content.strip()

    def extract_section(self, pattern, text):
        """パターンマッチングでセクションを抽出"""
        match = re.search(pattern, text, re.DOTALL | re.IGNORECASE)
        if match:
            return self.normalize_section_text(match.group(1))
        return ""

    def choose_target_cell(self, label_cell, sheet):
        """ラベルセルから書き込み先のターゲットセルを選択"""
        label_row = label_cell.row
        label_col = label_cell.column
        candidates = []
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            if min_row >= label_row + 1 and min_row <= label_row + 12:
                width = max_col - min_col + 1
                height = max_row - min_row + 1
                if width >= 6:
                    if min_col <= label_col <= max_col or min_col > label_col:
                        candidates.append(
                            (min_row, -(width * height), min_col, str(merged_range))
                        )
        if candidates:
            candidates.sort()
            selected = candidates[0][3]
            min_col, min_row, _, _ = range_boundaries(selected)
            return f"{get_column_letter(min_col)}{min_row}"
        return f"{get_column_letter(label_col + 1)}{label_row}"

    def test_template_exists(self, template_path):
        """テンプレートファイルが存在することを確認"""
        assert template_path.exists(), f"Template file not found: {template_path}"

    def test_template_structure(self, template_path):
        """テンプレートの基本構造を確認"""
        workbook = openpyxl.load_workbook(template_path)
        sheet = workbook.active

        assert sheet.title == "創業計画書", "Sheet name should be '創業計画書'"
        assert sheet.max_row > 0, "Sheet should have rows"
        assert sheet.max_column > 0, "Sheet should have columns"

    def test_find_motivation_label(self, template_path):
        """「創業の動機」ラベルが見つかることを確認"""
        workbook = openpyxl.load_workbook(template_path)
        sheet = workbook.active

        found = False
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "創業の動機" in cell.value:
                    found = True
                    assert cell.coordinate == "B6", (
                        f"Expected B6, but found at {cell.coordinate}"
                    )
                    break
            if found:
                break

        assert found, "Label '創業の動機' not found in template"

    def test_extract_motivation_section(self, sample_plan_text):
        """「創業の動機」セクションが正しく抽出できることを確認"""
        heading_order = [
            ("motivation", "創業の動機"),
            ("background", "経営者の略歴等"),
        ]

        sections = {}
        for idx, (key, label) in enumerate(heading_order):
            next_labels = [next_label for _, next_label in heading_order[idx + 1 :]]
            pattern = (
                self.build_section_pattern(label, next_labels)
                if next_labels
                else self.build_section_pattern(label, [])
            )
            sections[key] = pattern

        content = self.extract_section(sections["motivation"], sample_plan_text)

        assert content != "", "Extracted content should not be empty"
        assert "大学卒業後" in content, "Content should contain '大学卒業後'"
        assert "IT企業" in content, "Content should contain 'IT企業'"
        assert "受託開発事業を立ち上げることを決意しました" in content

    def test_write_motivation_to_excel(self, template_path, sample_plan_text):
        """「創業の動機」がExcelの正しいセル(B7)に書き込まれることを確認"""
        workbook = openpyxl.load_workbook(template_path)
        sheet = workbook.active

        # ラベルセルを探す
        label_cell = None
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "創業の動機" in cell.value:
                    label_cell = cell
                    break
            if label_cell:
                break

        assert label_cell is not None, "Label cell not found"

        # ターゲットセルを特定
        target_cell_addr = self.choose_target_cell(label_cell, sheet)
        assert target_cell_addr == "B7", f"Expected B7, but got {target_cell_addr}"

        # セクションを抽出
        heading_order = [
            ("motivation", "創業の動機"),
            ("background", "経営者の略歴等"),
        ]
        sections = {}
        for idx, (key, label) in enumerate(heading_order):
            next_labels = [next_label for _, next_label in heading_order[idx + 1 :]]
            pattern = (
                self.build_section_pattern(label, next_labels)
                if next_labels
                else self.build_section_pattern(label, [])
            )
            sections[key] = pattern

        content = self.extract_section(sections["motivation"], sample_plan_text)

        # セルに書き込み
        sheet[target_cell_addr].value = content

        # 一時ファイルに保存して検証
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)
            workbook.save(tmp_path)

        # 保存されたファイルを読み込んで検証
        saved_workbook = openpyxl.load_workbook(tmp_path)
        saved_sheet = saved_workbook.active
        saved_value = saved_sheet["B7"].value

        assert saved_value is not None, "B7 cell should not be None"
        assert "大学卒業後" in saved_value, "Saved content should contain '大学卒業後'"
        assert "IT企業" in saved_value, "Saved content should contain 'IT企業'"

        # クリーンアップ
        tmp_path.unlink()

    def test_motivation_section_end_to_end(self, template_path, sample_plan_text):
        """創業の動機セクションのエンドツーエンドテスト"""
        workbook = openpyxl.load_workbook(template_path)
        sheet = workbook.active

        # 1. ラベル検出
        label_to_key = {"創業の動機": "motivation"}
        label_cells = {}
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    value = cell.value.strip()
                    for label, key in label_to_key.items():
                        if label in value:
                            label_cells[key] = cell

        assert "motivation" in label_cells, "Label '創業の動機' should be found"

        # 2. ターゲットセル特定
        mapping = {}
        for key, cell in label_cells.items():
            mapping[key] = self.choose_target_cell(cell, sheet)

        assert mapping["motivation"] == "B7", "Target cell should be B7"

        # 3. コンテンツ抽出
        heading_order = [("motivation", "創業の動機"), ("background", "経営者の略歴等")]
        sections = {}
        for idx, (key, label) in enumerate(heading_order):
            next_labels = [next_label for _, next_label in heading_order[idx + 1 :]]
            pattern = (
                self.build_section_pattern(label, next_labels)
                if next_labels
                else self.build_section_pattern(label, [])
            )
            sections[key] = pattern

        content = self.extract_section(sections["motivation"], sample_plan_text)
        assert content, "Extracted content should not be empty"

        # 4. 書き込み
        cell_addr = mapping["motivation"]
        sheet[cell_addr].value = content

        # 5. 検証
        assert sheet["B7"].value is not None, "B7 should have content"
        assert "大学卒業後" in sheet["B7"].value
        assert "受託開発事業を立ち上げることを決意しました" in sheet["B7"].value
